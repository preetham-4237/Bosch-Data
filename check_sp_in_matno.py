"""
check_sp_in_matno.py
====================
Checks if Column C (Mat.No.(QM) SP — spare part number) values exist
in Column A (Mat.No. (QM) — product number) in BIQIC_Claims_SP+FC.xlsx.

Outputs a CSV with all rows + a 'found_in_col_a' boolean column.

Usage:
    python check_sp_in_matno.py
"""

import csv
from pathlib import Path

import openpyxl

EXCEL_FILE = Path(__file__).parent.parent / "BIQIC_Claims_SP+FC.xlsx"
OUTPUT_FILE = Path(__file__).parent / "sp_in_matno_check.csv"

print(f"Reading {EXCEL_FILE} ...")
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
ws = wb["Sheet1"]

# --- Pass 1: collect all unique Column A values ---
col_a_values = set()
total_rows = 0

for row in ws.iter_rows(min_row=2, values_only=True):
    total_rows += 1
    val_a = row[0]  # Column A
    if val_a is not None:
        col_a_values.add(str(val_a).strip())

print(f"Total data rows: {total_rows}")
print(f"Unique Column A (Mat.No. QM) values: {len(col_a_values)}")

# --- Pass 2: check Column C against Column A set, write CSV ---
wb.close()
wb = openpyxl.load_workbook(EXCEL_FILE, read_only=True)
ws = wb["Sheet1"]

found_count = 0
not_found_count = 0

with open(OUTPUT_FILE, "w", newline="", encoding="utf-8") as f:
    writer = csv.writer(f)

    # Header
    writer.writerow([
        "mat_no_qm", "mat_no_qm_desc",
        "mat_no_sp", "mat_no_sp_desc",
        "fault_category_code", "fault_category",
        "fault_code_1", "fault_code_1_desc",
        "fault_code_2", "fault_code_2_desc",
        "warranty_cases_cntr",
        "sp_found_in_col_a"
    ])

    for row in ws.iter_rows(min_row=2, values_only=True):
        val_c = row[2]  # Column C
        sp_found = False
        if val_c is not None:
            sp_found = str(val_c).strip() in col_a_values

        if sp_found:
            found_count += 1
        else:
            not_found_count += 1

        writer.writerow([
            row[0], row[1],   # A, B
            row[2], row[3],   # C, D
            row[4], row[5],   # E, F
            row[6], row[7],   # G, H
            row[8], row[9],   # I, J
            row[10],          # K
            sp_found
        ])

wb.close()

print(f"\nResults:")
print(f"  Total rows:         {total_rows}")
print(f"  SP found in Col A:  {found_count}")
print(f"  SP NOT in Col A:    {not_found_count}")
print(f"  Match rate:         {100 * found_count / total_rows:.1f}%")
print(f"\nOutput written to: {OUTPUT_FILE}")